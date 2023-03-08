# %% [markdown]
#    ># SM-S919O DM3 EUR Auto Cal Spec 변경 PGM

# %% [markdown]
#   <font size ="2">
#
#   === Quadra ===
#   - V1
#     - 1st Release
#     - RX Freq Cal Default 반영 시 수정전 파일의 주파수 자리수가 부족할 때 오류 발생
#       - Value list 를 삭제 후 Freq list 수대로 0으로 자리수 생성 후 값 변경하도록 수정
#     - RX_FREQ_CAL_EN=1 검색 수 아래로 2번째 줄에서 Freq list 얻어오도록 했으나, 공백 있을 때 오류 발생
#       - 수정 필요
#
#   </font>

# %%
import os
import glob

import ttkbootstrap as ttkbst
from ttkbootstrap.constants import *
import tkinter as tk
import tkinter.scrolledtext as st

import threading

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.numbers import builtin_format_code

import pandas as pd

import Common_function as func
import LSI_get_data as getdata

# %%
font_style = Font(
    name="Calibri",
    size=10,
    bold=False,
    italic=False,
    vertAlign=None,  # 첨자
    underline="none",  # 밑줄
    strike=False,  # 취소선
    color="00000000",  # 블랙, # 00FF0000 Red, # 000000FF Blue
)

# %%
def Common_save_Excel(filename, tab1, tab2):
    # Save Data to Excel
    Tabname = filename.replace("Excel_", "")
    Tabname = f"{os.path.splitext(Tabname)[0]}"
    with pd.ExcelWriter(filename) as writer:
        tab1.to_excel(writer, sheet_name=f"{Tabname}_Mean")
        tab2.to_excel(writer, sheet_name=f"{Tabname}_Data")


def WB_Format(filename):
    wb = load_workbook(filename)
    ws = wb.sheetnames
    for sheet in ws:
        col_max = wb[sheet].max_column
        row_max = wb[sheet].max_row
        for i in range(2, row_max + 1, 1):
            for j in range(3, col_max + 1, 1):
                wb[sheet].cell(row=i, column=j).font = font_style
                wb[sheet].cell(row=i, column=j).alignment = Alignment(horizontal="right")
                wb[sheet].cell(row=i, column=j).number_format = "#,##0.0"
                # wb[sheet].cell(row=i, column=j).number_format = builtin_format_code(2)
    wb.save(filename)


# %%
Win_GUI = ttkbst.Window(title="S23 Quadra DM3 EUR Auto Cal Spec PGM V1.0", themename="cosmo")
Win_GUI.attributes("-topmost", True)
Win_GUI.geometry("1410x565")  # py : 1407x560 ipynb : 1635x670
# Win_GUI.option_add("*Font", "Consolas 10")


def change_theme():
    themename = Win_GUI.getvar("themename")
    Win_GUI.style.theme_use(themename)


themes = Win_GUI.style.theme_names()

Left_frame = ttkbst.Frame(Win_GUI)
Left_frame.place(x=0, y=0, width=625, height=565)

# 리스트 프레임
list_frame = ttkbst.Frame(Left_frame)
list_frame.place(x=5, y=5, width=615, height=130)


scrollbar = ttkbst.Scrollbar(list_frame, orient="vertical")
scrollbar.place(x=590, y=45, width=25, height=80)

list_file = tk.Listbox(list_frame, height=5, yscrollcommand=scrollbar.set)
list_file.place(x=0, y=45, width=590, height=80)

# %%
Scrolled_txt_frame = ttkbst.Frame(Win_GUI)
Scrolled_txt_frame.place(x=625, y=5, width=780, height=555)

text_area = st.ScrolledText(Scrolled_txt_frame, font=("Consolas", 9))
text_area.place(x=5, y=5, width=770, height=505)

right_run_frame = ttkbst.Frame(Scrolled_txt_frame)
right_run_frame.place(x=5, y=510, width=780, height=55)

Author = ttkbst.Label(right_run_frame, text="dongsub.roh@samsung.com", anchor="w")
Author.place(x=5, y=5, width=200, height=45)

# p_var = tk.DoubleVar()
# progressbar = ttkbst.CTkProgressBar(right_run_frame, variable=p_var, fg_color="#dbdbdb", corner_radius=5)
# progressbar.place(x=0, y=5, width=770, height=45)

# %%
# 저장 경로 프레임
path_frame = ttkbst.Labelframe(Left_frame, text=" Select Files ", bootstyle=PRIMARY)
path_frame.place(x=5, y=135, width=615, height=85)

spc_label = ttkbst.Label(path_frame, text="SPC 파일")
spc_label.place(x=10, y=15, width=60, height=30)

txt_spc_path = ttkbst.Entry(path_frame)
# spc 파일 경로 사전입력
txt_spc_path.insert(
    0,
    "D:/DATA/Project_DATA/@_S23/LSI/TOOLS/3. DASEUL/제조사양서/SM-S919O_OPEN_CALIBRATION_Ver_3.1.632.0T3.spc.dec",
)
txt_spc_path.place(x=70, y=15, width=425, height=30)

btn_xml_path = ttkbst.Button(
    path_frame,
    text="SPC 선택 (F2)",
    command=lambda: [func.browse_spc_path(txt_spc_path, text_area)],
)
btn_xml_path.place(x=505, y=15, width=100, height=30)

Win_GUI.bind("<F2>", lambda event: [func.browse_spc_path(txt_spc_path, text_area)])

# %%
# 옵션 선택 frame
radio_Btn_frame = ttkbst.Labelframe(Left_frame, text=" Select Options ", bootstyle=PRIMARY)
radio_Btn_frame.place(x=5, y=225, width=615, height=100)

Option_var = ttkbst.IntVar()

btn_Option1 = ttkbst.Radiobutton(radio_Btn_frame, text="Cal Spec 조정", value=1, variable=Option_var)
btn_Option1.place(x=10, y=10, width=100, height=30)

btn_Option2 = ttkbst.Radiobutton(radio_Btn_frame, text="Cal 산포 적용", value=2, variable=Option_var)
btn_Option2.place(x=120, y=10, width=100, height=30)
btn_Option2.invoke()

btn_Option3 = ttkbst.Radiobutton(radio_Btn_frame, text="MTM Default Cal Data", value=3, variable=Option_var)
btn_Option3.place(x=230, y=10, width=140, height=30)

Save_data_var = ttkbst.BooleanVar()

chkbox = ttkbst.Checkbutton(radio_Btn_frame, text="Save Data to Excel", variable=Save_data_var)
chkbox.place(x=480, y=10, width=120, height=30)


def Mclick():
    global Select_op
    Select_op = "MTM"
    Option_var.set(3)


def Dclick():
    global Select_op
    Select_op = "Daseul"
    Option_var.set(2)


# Cal log 파일 선택
btn_add_file1 = ttkbst.Button(
    list_frame, text="Daseul log 추가 (F1)", command=lambda: [func.Common_daseul_log(list_file), Dclick()]
)
btn_add_file1.place(x=0, y=3, width=135, height=35)

btn_add_file2 = ttkbst.Button(
    list_frame, text="MTM log 추가 (F4)", command=lambda: [func.Common_mtm_log(list_file), Mclick()]
)
btn_add_file2.place(x=490, y=3, width=125, height=35)

# Cal log : log 폴더의 CSV 파일 자동 입력
for filename in glob.glob("C:\\DGS\\LOGS\\*.csv"):
    list_file.insert(tk.END, filename)

scrollbar.config(command=list_file.yview)

Win_GUI.bind("<F1>", lambda event: [func.Common_daseul_log(list_file), Dclick()])
Win_GUI.bind("<F4>", lambda event: [func.Common_mtm_log(list_file), Mclick()])

# %%
# Cal Spec frame
cal_spec_frame = ttkbst.Labelframe(Left_frame, text="Cal Spec", bootstyle=PRIMARY)
cal_spec_frame.place(x=5, y=330, width=615, height=185)

Cable_Spec_label = ttkbst.Label(cal_spec_frame, text="Cable Check", anchor="e")
Cable_Spec_label.place(x=10, y=10, width=70, height=25)
Cable_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
Cable_Spec_var.place(x=85, y=10, width=40, height=25)
Cable_Spec_var.insert(END, "3")

GMSK_Spec_label = ttkbst.Label(cal_spec_frame, text="2G GMSK", anchor="e")
GMSK_Spec_label.place(x=10, y=45, width=70, height=25)
GMSK_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
GMSK_Spec_var.place(x=85, y=45, width=40, height=25)
GMSK_Spec_var.insert(END, "5")

EPSK_Spec_label = ttkbst.Label(cal_spec_frame, text="2G EPSK", anchor="e")
EPSK_Spec_label.place(x=10, y=80, width=70, height=25)
EPSK_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
EPSK_Spec_var.place(x=85, y=80, width=40, height=25)
EPSK_Spec_var.insert(END, "5")

FBRX_Spec_label = ttkbst.Label(cal_spec_frame, text="NR FBRX", anchor="e")
FBRX_Spec_label.place(x=140, y=10, width=50, height=25)
FBRX_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
FBRX_Spec_var.place(x=195, y=10, width=40, height=25)
FBRX_Spec_var.insert(END, "5")

FBRX_3G_Spec_label = ttkbst.Label(cal_spec_frame, text="3G FBRX", anchor="e")
FBRX_3G_Spec_label.place(x=140, y=45, width=50, height=25)
FBRX_3G_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
FBRX_3G_Spec_var.place(x=195, y=45, width=40, height=25)
FBRX_3G_Spec_var.insert(END, "3")

FBRX_2G_Spec_label = ttkbst.Label(cal_spec_frame, text="2G FBRX", anchor="e")
FBRX_2G_Spec_label.place(x=140, y=80, width=50, height=25)
FBRX_2G_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
FBRX_2G_Spec_var.place(x=195, y=80, width=40, height=25)
FBRX_2G_Spec_var.insert(END, "3")

RX_Gain_label = ttkbst.Label(cal_spec_frame, text="NR RX Gain", anchor="e")
RX_Gain_label.place(x=240, y=10, width=70, height=25)
RX_Gain_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
RX_Gain_Spec_var.place(x=315, y=10, width=40, height=25)
RX_Gain_Spec_var.insert(END, "5")

RX_Gain_3G_label = ttkbst.Label(cal_spec_frame, text="3G RX Gain", anchor="e")
RX_Gain_3G_label.place(x=240, y=45, width=70, height=25)
RX_Gain_3G_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
RX_Gain_3G_Spec_var.place(x=315, y=45, width=40, height=25)
RX_Gain_3G_Spec_var.insert(END, "5")

RX_Gain_2G_label = ttkbst.Label(cal_spec_frame, text="2G RX Gain", anchor="e")
RX_Gain_2G_label.place(x=240, y=80, width=70, height=25)
RX_Gain_2G_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
RX_Gain_2G_Spec_var.place(x=315, y=80, width=40, height=25)
RX_Gain_2G_Spec_var.insert(END, "5")

APT_Spec_label = ttkbst.Label(cal_spec_frame, text="APT_Spec", anchor="e")
APT_Spec_label.place(x=360, y=10, width=70, height=25)
APT_Spec_var = ttkbst.Entry(cal_spec_frame, justify="right")
APT_Spec_var.place(x=435, y=10, width=40, height=25)
APT_Spec_var.insert(END, "0.5")

ET_Psat_label = ttkbst.Label(cal_spec_frame, text="ET_Psat", anchor="e")
ET_Psat_label.place(x=480, y=10, width=70, height=25)
ET_Psat_var = ttkbst.Entry(cal_spec_frame, justify="right")
ET_Psat_var.place(x=555, y=10, width=40, height=25)
ET_Psat_var.insert(END, "3")

ET_Pgain_label = ttkbst.Label(cal_spec_frame, text="ET_Pgain", anchor="e")
ET_Pgain_label.place(x=480, y=45, width=70, height=25)
ET_Pgain_var = ttkbst.Entry(cal_spec_frame, justify="right")
ET_Pgain_var.place(x=555, y=45, width=40, height=25)
ET_Pgain_var.insert(END, "3")

ET_Freq_label = ttkbst.Label(cal_spec_frame, text="ET_Freq", anchor="e")
ET_Freq_label.place(x=480, y=80, width=70, height=25)
ET_Freq_var = ttkbst.Entry(cal_spec_frame, justify="right")
ET_Freq_var.place(x=555, y=80, width=40, height=25)
ET_Freq_var.insert(END, "3")

ET_Power_label = ttkbst.Label(cal_spec_frame, text="ET_Power", anchor="e")
ET_Power_label.place(x=480, y=115, width=70, height=25)
ET_Power_var = ttkbst.Entry(cal_spec_frame, justify="right")
ET_Power_var.place(x=555, y=115, width=40, height=25)
ET_Power_var.insert(END, "3")

# %%
# 실행 프레임
left_run_frame = ttkbst.Frame(Left_frame)
left_run_frame.place(x=5, y=520, width=615, height=40)

theme_options = tk.Menubutton(left_run_frame, text="Select a theme")
menu = tk.Menu(theme_options)

for t in themes:
    menu.add_radiobutton(label=t, variable="themename", command=change_theme)

theme_options["menu"] = menu
theme_options.place(x=0, y=5, width=100, height=30)

btn_start = ttkbst.Button(
    left_run_frame,
    text="시작 (F5)",
    command=lambda: [
        threading.Thread(
            target=getdata.start,
            args=(
                list_file,
                Option_var,
                txt_spc_path,
                Select_op,
                Cable_Spec_var,
                RX_Gain_Spec_var,
                FBRX_Spec_var,
                APT_Spec_var,
                ET_Psat_var,
                ET_Pgain_var,
                ET_Freq_var,
                ET_Power_var,
                RX_Gain_3G_Spec_var,
                FBRX_3G_Spec_var,
                RX_Gain_2G_Spec_var,
                GMSK_Spec_var,
                EPSK_Spec_var,
                Save_data_var,
                text_area,
            ),
        ).start()
    ],
)
btn_start.place(x=415, y=0, width=200, height=40)

Win_GUI.bind(
    "<F5>",
    lambda event: [
        threading.Thread(
            target=getdata.start,
            args=(
                list_file,
                Option_var,
                txt_spc_path,
                Select_op,
                Cable_Spec_var,
                RX_Gain_Spec_var,
                FBRX_Spec_var,
                APT_Spec_var,
                ET_Psat_var,
                ET_Pgain_var,
                ET_Freq_var,
                ET_Power_var,
                RX_Gain_3G_Spec_var,
                FBRX_3G_Spec_var,
                RX_Gain_2G_Spec_var,
                GMSK_Spec_var,
                EPSK_Spec_var,
                Save_data_var,
                text_area,
            ),
        ).start()
    ],
)

Win_GUI.resizable(False, False)
Win_GUI.mainloop()
