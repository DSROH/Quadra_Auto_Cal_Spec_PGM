import os
import tkinter as tk
from tkinter import filedialog

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.numbers import builtin_format_code


def Common_save_Excel(filename, tab1, tab2):
    # Save Data to Excel
    Tabname = filename.replace("Excel_", "")
    Tabname = f"{os.path.splitext(Tabname)[0]}"
    with pd.ExcelWriter(filename) as writer:
        tab1.to_excel(writer, sheet_name=f"{Tabname}_Mean")
        tab2.to_excel(writer, sheet_name=f"{Tabname}_Data")


def WB_Format(filename, i, j, k):
    font_style = Font(
        name="Calibri",
        size=10,
        bold=False,
        italic=False,
        vertAlign=None,  # 첨자
        underline="none",  # 밑줄
        strike=False,  # 취소선
        color="00000000",  # 블랙, # 00FF0000 Red, # 000000FF Blue
    )
    wb = load_workbook(filename)
    ws = wb.sheetnames
    for sheet in ws:
        col_max = wb[sheet].max_column
        row_max = wb[sheet].max_row
        for row_c in range(i, row_max + 1, 1):
            for col_c in range(j, col_max + 1, 1):
                wb[sheet].cell(row=row_c, column=col_c).font = font_style
                wb[sheet].cell(row=row_c, column=col_c).alignment = Alignment(horizontal="right")
                # wb[sheet].cell(row=row_c, column=col_c).number_format = "#,##0.0"
                wb[sheet].cell(row=row_c, column=col_c).number_format = builtin_format_code(k)
    wb.save(filename)


def Common_daseul_log(list_file):
    list_file.delete(0, tk.END)
    files = filedialog.askopenfilenames(
        title="Cal log 파일을 선택하세요",
        filetypes=(("CSV 파일", "*.csv"), ("모든 파일", "*.*")),
        initialdir=r"C:\DGS\LOGS",
    )
    # 사용자가 선택한 파일 목록
    for file in files:
        list_file.insert(tk.END, file)


def Common_mtm_log(list_file):
    list_file.delete(0, tk.END)
    files = filedialog.askopenfilenames(
        title="MTM log 파일을 선택하세요",
        filetypes=(("CSV 파일", "*.csv"), ("모든 파일", "*.*")),
        initialdir=(r"D:\\DATA\\Project_DATA\\@_S23\\LSI\\TOOLS\\5_MTM_Calibration\\SHAANON_MTM_3.71.01_DM3_20220721"),
    )
    # 사용자가 선택한 파일 목록
    for file in files:
        list_file.insert(tk.END, file)


def browse_spc_path(txt_spc_path, text_area):
    spc_file_name = filedialog.askopenfilename(
        title="SPC 파일을 선택하세요",
        filetypes=(("SPC 파일", "*.dec"), ("모든 파일", "*.*")),
        initialdir=r"D:\DATA\TOOLS\@_Python\Auto_Cal_Spec",
    )

    if spc_file_name == "":  # 사용자가 취소를 누를 때
        text_area.insert(tk.END, "폴더 선택 취소\n")
        text_area.see(tk.END)
        return
    # print(folder_selected)
    txt_spc_path.delete(0, tk.END)
    txt_spc_path.insert(0, spc_file_name)
